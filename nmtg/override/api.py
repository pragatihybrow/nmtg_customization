import frappe
from frappe.model.naming import make_autoname
from frappe.utils import today
from frappe import _
import json
import re
from frappe.utils import escape_html, get_url
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from frappe.utils.file_manager import save_file
import base64
import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from frappe.utils.file_manager import save_file




REMARKS_KEY = "__remarks__"


STANDARD_ITEM_FIELDS = {
    "item_code", "qty", "uom", "warehouse", "rate", "lead_time_days",
    "material_request", "material_request_item"
}

HEADER_CUSTOM_FIELDS = [
    "custom_purchase_type",
    "custom_other_purchase_type",
    "custom_rfq_nature",
    "custom_price_validity",
    "custom_warranty__guarantee",
    "custom_payment_terms",
    "custom_credit_days",
    "custom_advance_required",
    "custom_advance_required_amount",
    "custom_payment_milestone",
    "custom_retention__security_deposit_",
    "custom_retention__security_deposit_amount",
    "custom_supplier_contact_person",
    "custom_certificate_commitment",
    "custom_supplier_email",
    "custom_supplier_mobile",
    "custom_general_terms_acceptance",
    "custom_no_hidden_charges_declaration",
    "custom_quotation_correctness_declaration",
    "custom_authorized_person_name",
    "custom_designation",
    "custom__contact_number",
]

@frappe.whitelist(allow_guest=True)
def submit_supplier_quotation(data):
    if isinstance(data, str):
        data = json.loads(data)

    required = ['supplier', 'company', 'valid_till', 'rfq', 'items']
    for field in required:
        if not data.get(field):
            frappe.throw(f"Missing required field: {field}")

    rfq = frappe.get_doc("Request for Quotation", data['rfq'])
    supplier_names = [s.supplier for s in rfq.suppliers]
    if data['supplier'] not in supplier_names:
        frappe.throw("Supplier not authorized for this RFQ")

    existing = frappe.db.exists("Supplier Quotation", {
        "supplier": data['supplier'],
        "rfq": data['rfq']
    })
    if existing:
        frappe.throw(
            f"A quotation from {data['supplier']} for {data['rfq']} already exists: {existing}"
        )

    transaction_date = frappe.utils.today()

    if frappe.utils.getdate(data['valid_till']) < frappe.utils.getdate(transaction_date):
        frappe.throw(
            _("Valid till date cannot be before {0} (today). Please choose a later date and resubmit.")
            .format(frappe.utils.formatdate(transaction_date))
        )

    doc = frappe.new_doc("Supplier Quotation")
    doc.supplier = data['supplier']
    doc.company = data['company']
    doc.transaction_date = transaction_date
    doc.custom_submission_date_and_time = frappe.utils.now_datetime()
    doc.valid_till = data['valid_till']
    doc.rfq = data['rfq']

    if data.get('terms'):
        doc.terms = data['terms']
    if data.get('payment_terms_template'):
        doc.payment_terms_template = data['payment_terms_template']

    # ---- tax category / template ----
    if data.get('tax_category'):
        doc.tax_category = data['tax_category']
    if data.get('taxes_and_charges'):
        doc.taxes_and_charges = data['taxes_and_charges']

    # ---- header-level custom fields ----
    for fieldname in HEADER_CUSTOM_FIELDS:
        if fieldname in data:
            doc.set(fieldname, data[fieldname])

    # ---- items, including per-purchase-type custom fields ----
    for item in data['items']:
        row = {
            "item_code": item['item_code'],
            "qty": item['qty'],
            "uom": item.get('uom', 'Nos'),
            "stock_uom": item.get('uom', 'Nos'),
            "warehouse": item.get('warehouse') or rfq.set_warehouse or '',
            "rate": item['rate'],
            "lead_time_days": item.get('lead_time_days', 0),
            "material_request": item.get('material_request', ''),
            "material_request_item": item.get('material_request_item', ''),
            "request_for_quotation": data['rfq'],
        }
        for key, value in item.items():
            if key not in STANDARD_ITEM_FIELDS:
                row[key] = value

        doc.append("items", row)

    # ---- taxes and charges table ----
    for tax in data.get('taxes', []):
        doc.append("taxes", {
            "category": tax.get('category', 'Total'),
            "add_deduct_tax": tax.get('add_deduct_tax', 'Add'),
            "charge_type": tax.get('charge_type', 'On Net Total'),
            "row_id": tax.get('row_id', ''),
            "account_head": tax.get('account_head', ''),
            "description": tax.get('description', ''),
            "rate": tax.get('rate', 0),
            "tax_amount": tax.get('tax_amount', 0),
        })

    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": doc.name, "status": "created"}

@frappe.whitelist(allow_guest=True)
def get_rfq_for_supplier(rfq, supplier):
    doc = frappe.get_doc("Request for Quotation", rfq)
    return {
        "rfq": {
            "name": doc.name,
            "company": doc.company,
            "transaction_date": str(doc.transaction_date),
            "schedule_date": str(doc.schedule_date),
            "supplier_name": supplier,
            # "set_warehouse": doc.set_warehouse,
            "custom_purchase_type": getattr(doc, "custom_purchase_type", ""),
            "custom_rfq_nature": getattr(doc, "custom_rfq_nature", ""),
        },
        "items": [{
            "idx": it.idx,
            "item_code": it.item_code,
            "item_name": it.item_name,
            "item_group": it.item_group,
            "qty": it.qty,
            "uom": it.uom,
            "warehouse": it.warehouse,
            "material_request": it.material_request,
            "material_request_item": it.material_request_item,
            "custom_tds_attachment": it.custom_tds_attachment,
            "image": it.image
        } for it in doc.items]
    }




@frappe.whitelist(allow_guest=True)
def get_purchase_taxes_templates(company=None):
    """Lightweight list for the guest-page autocomplete."""
    filters = {}
    if company:
        filters["company"] = company
    return frappe.get_all(
        "Purchase Taxes and Charges Template",
        filters=filters,
        fields=["name", "tax_category", "is_default"],
        order_by="is_default desc, name asc",
        limit_page_length=0,
    )


@frappe.whitelist(allow_guest=True)
def get_taxes_template_details(template):
    """Returns the template's tax_category + its taxes rows, for populating the guest form."""
    doc = frappe.get_doc("Purchase Taxes and Charges Template", template)
    return {
        "tax_category": doc.tax_category,
        "taxes": [
            {
                "category": t.category,
                "add_deduct_tax": t.add_deduct_tax,
                "charge_type": t.charge_type,
                "account_head": t.account_head,
                "description": t.description,
                "rate": t.rate,
            }
            for t in doc.taxes
        ],
    }
    

@frappe.whitelist()
def create_heat_number(po, row_name):

    row = frappe.get_value(
        "Purchase Receipt Item",
        row_name,
        ["qty", "custom_qty_in_no", "custom_nmtg_heat_number", "custom_single_heat_number"],
        as_dict=True
    )

    if not row:
        frappe.throw("Row not found")

    if row.custom_nmtg_heat_number:
        frappe.throw("Heat Number already generated")

    qty = int(row.custom_qty_in_no or 0)  # ✅ use custom_qty_in_no instead of qty

    if qty <= 0:
        frappe.throw("Please set a valid Qty In No before generating Heat Number.")

    fiscal_year = frappe.db.get_value(
        "Fiscal Year",
        {
            "year_start_date": ["<=", today()],
            "year_end_date": [">=", today()],
            "disabled": 0
        },
        "name"
    )

    if not fiscal_year:
        frappe.throw("No active Fiscal Year found")

    year_initial = frappe.db.get_value(
        "Fiscal Year",
        fiscal_year,
        "custom__year_initial"
    )

    if not year_initial:
        frappe.throw(f"Year Initial missing in Fiscal Year {fiscal_year}")

    prefix = f"N{year_initial}"

    if row.custom_single_heat_number:
        single = make_autoname(f"{prefix}.###")
        num = int(single.replace(prefix, "").replace(".", ""))
        heat_number = f"{prefix}{str(num).zfill(3)}"
    else:
        first = make_autoname(f"{prefix}.###")
        start = int(first.replace(prefix, "").replace(".", ""))
        end = start

        for _ in range(qty - 1):
            nxt = make_autoname(f"{prefix}.###")
            end = int(nxt.replace(prefix, "").replace(".", ""))

        heat_number = (
            f"{prefix}{str(start).zfill(3)}"
            f" - "
            f"{prefix}{str(end).zfill(3)}"
        )

    frappe.db.set_value(
        "Purchase Receipt Item",
        row_name,
        "custom_nmtg_heat_number",
        heat_number,
        update_modified=False
    )

    return heat_number


@frappe.whitelist()
def make_quality_inspections_custom(doctype, docname, company, items, inspection_type="Incoming"):
    import json

    if isinstance(items, str):
        items = json.loads(items)

    valid_types = set(frappe.db.sql_list(
        "SELECT name FROM `tabQC Testing Type`"
    ))

    # Get supplier from GRN (Purchase Receipt)
    grn_supplier = frappe.db.get_value(doctype, docname, "supplier")

    qi_names = []

    for item in items:
        qi = frappe.new_doc("Quality Inspection")
        qi.inspection_type = inspection_type
        qi.reference_type = doctype
        qi.reference_name = docname
        qi.item_code = item.get("item_code")
        qi.item_name = item.get("item_name")
        qi.sample_size = item.get("sample_size") or item.get("qty") or 1
        qi.inspected_by = frappe.session.user
        qi.report_date = frappe.utils.today()
        qi.company = company
        qi.child_row_reference = item.get("child_row_reference", "")

        # Lab Name
        qi.custom_supplier = item.get("supplier", "")

        # Actual GRN Supplier Name
        qi.custom_supplier_name = grn_supplier

        qi.custom_nmtg_heat_number = item.get("nmtg_heat_number", "")
        qi.custom_vendor_heat_number = item.get("custom_vendor_heat_number", "")
        qi.custom_mill_tc = item.get("custom_mill_tc", "")

        testing_value = item.get("testing_value", "[]")

        try:
            testing_types = (
                json.loads(testing_value)
                if isinstance(testing_value, str)
                else (testing_value or [])
            )
        except Exception:
            testing_types = []

        for tt in testing_types:
            tt = (tt or "").strip()
            if tt and tt in valid_types:
                qi.append("custom_testing_type", {
                    "testing_type": tt
                })

        qi.insert(ignore_permissions=True)
        qi_names.append(qi.name)

    frappe.db.commit()
    return qi_names
    


@frappe.whitelist()
def update_qc_testing_type(row_name, testing_value):
    frappe.db.set_value(
        "Supplier Selection For QC",
        row_name,
        "testing_value",
        testing_value,
        update_modified=False
    )
    frappe.db.commit()
    return True


@frappe.whitelist()
def get_or_create_qc_series(qi_names):
    import json
    from frappe.utils import today

    if isinstance(qi_names, str):
        qi_names = json.loads(qi_names)

    # Check if all docs already share a qc_series
    existing_series = set()
    for name in qi_names:
        val = frappe.db.get_value("Quality Inspection", name, "custom_qc_series")
        if val:
            existing_series.add(val)

    # All already assigned the same series — reuse it
    if len(existing_series) == 1:
        return existing_series.pop()

    # Generate new series using fiscal year initials
    fiscal_year = frappe.db.get_value(
        "Fiscal Year",
        {
            "year_start_date": ["<=", today()],
            "year_end_date":   [">=", today()],
            "disabled": 0
        },
        "name"
    )

    if not fiscal_year:
        frappe.throw("No active Fiscal Year found")

    # Fiscal year name is like "2026-2027", extract "26-27"
    fy_short = "-".join([part[-2:] for part in fiscal_year.split("-")])

    prefix = f"N-OL-{fy_short}-"

    # Get next sequence number
    series_key = f"{prefix}.####"
    next_val = frappe.model.naming.make_autoname(series_key)
    # next_val will be like "N-OL-26-27-0001"

    # Save to all docs in this group
    for name in qi_names:
        frappe.db.set_value(
            "Quality Inspection",
            name,
            "custom_qc_series",
            next_val,
            update_modified=False
        )

    frappe.db.commit()
    return next_val


def compute_row_status(doc, method=None):
	for row in doc.get("readings", []):
		if not row.custom_reading_details:
			continue
		new_status = _compute_row_status(row)
		if new_status:
			row.status = new_status


def _compute_row_status(row):
	try:
		details = json.loads(row.custom_reading_details or "{}")
	except Exception:
		details = {}

	accepted = 0
	rejected = 0

	for key, value in details.items():
		if key == REMARKS_KEY:
			continue

		if value in (None, ""):
			continue

		if row.manual_inspection:
			accepted += 1
			continue

		if not row.numeric:
			continue

		try:
			min_v = float(row.min_value)
			max_v = float(row.max_value)
		except (TypeError, ValueError):
			continue

		if min_v == 0 and max_v == 0:
			continue

		try:
			num = float(value)
		except (TypeError, ValueError):
			rejected += 1
			continue

		if min_v <= num <= max_v:
			accepted += 1
		else:
			rejected += 1

	if accepted == 0 and rejected == 0:
		return None
	if rejected == 0:
		return "Accepted"
	if accepted == 0:
		return "Rejected"
	return "Partially Accepted"


def update_parent_status(doc, method=None):
	"""
	Server-side source of truth for the parent `status` field.
	Must run AFTER compute_row_status, so it reads the freshly
	recomputed row statuses rather than stale ones.
	"""
	statuses = [(row.status or "").strip() for row in doc.get("readings", [])]
	statuses = [s for s in statuses if s]

	if not statuses:
		return

	if all(s == "Accepted" for s in statuses):
		doc.status = "Accepted"
	elif all(s == "Rejected" for s in statuses):
		doc.status = "Rejected"
	else:
		doc.status = "Partially Accepted"


def validate_heat_range_vs_sample_size(doc, method=None):
    if doc.workflow_state != "Draft":
        return

    heat_range = (doc.custom_nmtg_heat_number or "").strip()

    if not heat_range:
        frappe.throw(
            "Please enter NMTG Heat Number before sending for Internal QC Inspection."
        )

    parts = re.split(r"\s*-\s*", heat_range)

    if len(parts) != 2:
        frappe.throw(
            f"Invalid heat number format: {heat_range}. Example: NK272 - NK281"
        )

    m1 = re.match(r'^([A-Za-z]*)(\d+)$', parts[0].strip())
    m2 = re.match(r'^([A-Za-z]*)(\d+)$', parts[1].strip())

    if not m1 or not m2:
        frappe.throw("Invalid NMTG Heat Number format.")

    start_num = int(m1.group(2))
    end_num = int(m2.group(2))

    if end_num < start_num:
        frappe.throw("End heat number cannot be smaller than start heat number.")

    heat_count = (end_num - start_num) + 1
    sample_size = doc.sample_size or 0

    if heat_count != sample_size:
        prefix = m1.group(1)
        suggested_end = start_num + sample_size - 1
        frappe.throw(
            f"Heat Range Count ({heat_count}) does not match Sample Size ({sample_size}). "
            f"For Sample Size {sample_size}, the range should be exactly "
            f"{prefix}{start_num} - {prefix}{suggested_end}."
        )


def validate_quality_category_before_submit(doc, method):
    item_codes = list({d.item_code for d in doc.items if d.item_code})
    if not item_codes:
        return

    quality_required_map = dict(
        frappe.get_all(
            "Item",
            filters={"item_code": ["in", item_codes]},
            fields=["item_code", "custom_quality_category_required"],
            as_list=True,
        )
    )

    flagged_items = [
        row.item_code
        for row in doc.items
        if quality_required_map.get(row.item_code)
    ]

    if flagged_items and not doc.custom_quality_category:
        frappe.throw(
            _(
                "Quality Category is mandatory before submission because the "
                "following item(s) require it: {0}"
            ).format(", ".join(sorted(set(flagged_items))))
        )


@frappe.whitelist()
def custom_set_rejection_remark(name, remark):
    frappe.db.set_value("Supplier", name, "custom_rejection_remark", remark)
    frappe.db.commit()


@frappe.whitelist(allow_guest=True)
def get_supplier_request_type(supplier_name):
    if not supplier_name:
        return {}
    request_type = frappe.db.get_value("Supplier", supplier_name, "request_type")
    return {"request_type": request_type}


import json
import frappe


@frappe.whitelist(allow_guest=True)
def submit_supplier_registration():
    payload = json.loads(frappe.request.data)
    doc_data = payload.get("doc") or payload
    doc_data["doctype"] = "Supplier Registration Form"

    doc = frappe.get_doc(doc_data)
    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {"name": doc.name}


@frappe.whitelist(allow_guest=True)
def get_link_options():
    doctypes = ["Country", "Supplier", "Competency"]
    result = {}
    for doctype in doctypes:
        result[doctype] = frappe.get_all(
            doctype,
            pluck="name",
            order_by="name asc",
            limit_page_length=0,
        )
    return result


@frappe.whitelist(allow_guest=True)
def upload_supplier_file():
    """
    Guest-safe file upload for the public Supplier Registration Form.
    Bypasses the core /api/method/upload_file permission check
    (which requires create permission on File, which Guest lacks)
    while still validating that a real file was posted.
    """
    if "file" not in frappe.request.files:
        frappe.throw("No file was uploaded.")

    file = frappe.request.files["file"]
    filename = file.filename
    content = file.stream.read()

    if not content:
        frappe.throw("Uploaded file is empty.")

    saved = save_file(
        fname=filename,
        content=content,
        dt=None,          # not attached to a specific doc yet — form isn't saved until submit
        dn=None,
        is_private=1,
        ignore_permissions=True,
    )

    return {"file_url": saved.file_url, "file_name": saved.file_name}



@frappe.whitelist(allow_guest=True)
def upload_supplier_file():
    if "file" not in frappe.request.files:
        frappe.throw("No file was uploaded.")

    file = frappe.request.files["file"]
    filename = file.filename
    content = file.stream.read()

    if not content:
        frappe.throw("Uploaded file is empty.")

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": filename,
        "content": base64.b64encode(content).decode("utf-8"),
        "decode": True,
        "is_private": 1,
        "attached_to_doctype": None,
        "attached_to_name": None,
    })
    file_doc.insert(ignore_permissions=True)

    return {"file_url": file_doc.file_url, "file_name": file_doc.file_name}

@frappe.whitelist(allow_guest=True)
def save_supplier_audit(doc):
    if isinstance(doc, str):
        doc = json.loads(doc)

    if doc.get("doctype") != "Supplier Audit":
        frappe.throw("Invalid doctype.")

    audit = frappe.get_doc(doc)
    audit.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": audit.name}

def validate_lead(doc, method=None):
    if doc.workflow_state != "Qualified":
        return

    missing_fields = []

    if not doc.custom_customer_type:
        missing_fields.append(_("Customer Type"))

    if not doc.custom_industry_ct:
        missing_fields.append(_("Industry"))

    # if not doc.custom_application:
    #     missing_fields.append(_("Application"))

    if not doc.custom_product_group:
        missing_fields.append(_("Product Group"))

    if missing_fields:
        frappe.throw(
            _("The following fields are mandatory before changing the Workflow State to Qualified:<br><br><b>{0}</b>").format(
                "<br>".join(f"• {field}" for field in missing_fields)
            )
        )


@frappe.whitelist()
def create_or_update_lead_contact(row, lead):
    if isinstance(row, str):
        row = json.loads(row)
    row = frappe._dict(row)
    contact_name = row.get('custom_contact_ref')

    # Validate the stored ref is still linked to this lead
    if contact_name:
        linked = frappe.db.exists("Dynamic Link", {
            "parent": contact_name,
            "parenttype": "Contact",
            "link_doctype": "Lead",
            "link_name": lead
        })
        if not linked:
            contact_name = None

    # Dedup by email, but only if that contact is already linked to this lead
    if not contact_name and row.get('email_id'):
        existing = frappe.db.sql("""
            SELECT ce.parent
            FROM `tabContact Email` ce
            INNER JOIN `tabDynamic Link` dl
                ON dl.parent = ce.parent
                AND dl.parenttype = 'Contact'
                AND dl.link_doctype = 'Lead'
                AND dl.link_name = %s
            WHERE ce.email_id = %s
            LIMIT 1
        """, (lead, row.get('email_id')))
        if existing:
            contact_name = existing[0][0]

    if contact_name and frappe.db.exists("Contact", contact_name):
        contact = frappe.get_doc("Contact", contact_name)
    else:
        contact = frappe.new_doc("Contact")

    # ensure this Lead is linked, whether contact is new or reused
    already_linked = any(
        l.link_doctype == "Lead" and l.link_name == lead
        for l in contact.get('links', [])
    )
    if not already_linked:
        contact.append('links', {
            'link_doctype': 'Lead',
            'link_name': lead
        })

    contact.first_name = row.get('name1') or contact.first_name or "Contact"
    contact.designation = row.get('designation')

    if row.get('email_id'):
        # ensure no other row is marked primary before adding/updating this one
        for e in contact.get('email_ids', []):
            e.is_primary = 0

        existing_email = next(
            (e for e in contact.get('email_ids', []) if e.email_id == row.get('email_id')),
            None
        )
        if existing_email:
            existing_email.is_primary = 1
        else:
            contact.append('email_ids', {'email_id': row.get('email_id'), 'is_primary': 1})

    if row.get('contact_no'):
        for p in contact.get('phone_nos', []):
            p.is_primary_phone = 0

        existing_phone = next(
            (p for p in contact.get('phone_nos', []) if p.phone == row.get('contact_no')),
            None
        )
        if existing_phone:
            existing_phone.is_primary_phone = 1
        else:
            contact.append('phone_nos', {'phone': row.get('contact_no'), 'is_primary_phone': 1})

    if row.get('whatsapp_no'):
        contact.custom_whatsapp_no = row.get('whatsapp_no')

    if contact.is_new():
        contact.insert(ignore_permissions=True)
    else:
        contact.save(ignore_permissions=True)

    return contact.name




STAGE_FIELD_MAP = {
	"intro": "intro",
	"followup1": "followup_1",
	"followup2": "followup_2",
	"final": "final",
}


ATTACHMENT_FILE_URL = "/files/NMTG_Company_Profile.pdf"


def get_lead_recipients(doc):
	"""Primary contact (primary_contact=1) -> To; rest with an email -> CC.
	Falls back to the last row if no primary is marked."""
	contacts = doc.get("custom_contact_info") or []

	emails = []
	primary_email = None

	for row in contacts:
		if not row.email_id:
			continue
		if row.primary_contact and not primary_email:
			primary_email = row.email_id
		if row.email_id not in emails:
			emails.append(row.email_id)

	if not emails:
		return None, []

	to = primary_email or emails[-1]
	cc = [e for e in emails if e != to]

	return to, cc


def get_email_attachments():
	if not ATTACHMENT_FILE_URL:
		return []
	try:
		file_doc = frappe.get_doc("File", {"file_url": ATTACHMENT_FILE_URL})
		return [{"fname": file_doc.file_name, "fcontent": file_doc.get_content()}]
	except frappe.DoesNotExistError:
		frappe.log_error("NMTG stage email: attachment file not found", ATTACHMENT_FILE_URL)
		return []


@frappe.whitelist()
def send_lead_stage_email(lead, stage_key, subject, message):
	if stage_key not in STAGE_FIELD_MAP:
		frappe.throw(_("Invalid stage key: {0}").format(stage_key))

	if not frappe.has_permission("Lead", "write", lead):
		frappe.throw(_("Not permitted"), frappe.PermissionError)

	doc = frappe.get_doc("Lead", lead)

	to, cc = get_lead_recipients(doc)
	if not to:
		frappe.throw(_("No contact email addresses found to send to."))

	frappe.sendmail(
		recipients=[to],
		cc=cc,
		subject=subject,
		message=message,
		reference_doctype="Lead",
		reference_name=lead,
		attachments=get_email_attachments(),
	)

	field_key = STAGE_FIELD_MAP[stage_key]
	sent_on = frappe.utils.now()
	doc.db_set(f"custom_{field_key}_email_sent", 1, update_modified=False)
	doc.db_set(f"custom_{field_key}_email_sent_on", sent_on, update_modified=False)

	return {"sent_on": sent_on, "recipient": to, "cc": cc}


import json

import frappe
from frappe import _
from frappe.utils import escape_html, get_url


ITEM_FIELDS = [
	"request_no",
	"product_group",
	"customer_requirement__brief",
	"qty",
	"competitor_name",
	"competitor_model",
	"nmtg_model",
	"item_code",
	"item_name",
	"selection_type",
	"technical_status",
	"clarification_required",
	"drawing_approval_required",
	"drawing",
	"feasibility_review_required",
	"apqp_required",
	"apqp",
	"remark",
]

DETAIL_FIELDS = [
	"name",
	"reference_type",
	"opportunity_form",
	"customer",
	"enquiry_reference",
	"overall_technical_status",
	"lead_engineer",
	"email",
	"version",
	"general_technical_remarks",
]

@frappe.whitelist()
def send_technical_evaluation_questionary(technical_evaluation):
    """Send technical evaluation questionnaire email to the customer."""

    doc = frappe.get_doc("Technical Evaluation", technical_evaluation)

    if not doc.email:
        frappe.throw(_("Email is not set on {0}").format(doc.name))

    if not doc.questionary_for_technical_evaluation:
        frappe.throw(_("No questions found on {0} to send").format(doc.name))

    link = get_url(
        f"/questionary-for-technical-evaluation?technical_evaluation={doc.name}"
    )

    # Create mapping of Request No -> Item Name / Customer Requirement
    item_requirements = {
        item.request_no: item.item_name or ""
        for item in doc.items
        if item.request_no
    }

    rows_html = ""

    for row in doc.questionary_for_technical_evaluation:

        requirement = item_requirements.get(row.request_no, "")

        rows_html += f"""
            <tr>
                <td style="
                    padding:8px 10px;
                    border:1px solid #ddd;
                    vertical-align:top;
                ">
                    {escape_html(row.request_no or '')}
                </td>

                <td style="
                    padding:8px 10px;
                    border:1px solid #ddd;
                    vertical-align:top;
                ">
                    {escape_html(requirement)}
                </td>

                <td style="
                    padding:8px 10px;
                    border:1px solid #ddd;
                    vertical-align:top;
                ">
                    {escape_html(row.question or '')}
                </td>
            </tr>
        """

    message = f"""
        <div style="
            padding:0 0 16px 0;
            border-bottom:2px solid #000000;
            margin-bottom:20px;
        ">
            <h2 style="
                margin:0;
                font-family:Arial, Helvetica, sans-serif;
                color:#000000;
                font-size:20px;
                font-weight:700;
            ">
                NMTG Mechtrans Techniques Pvt. Ltd.
            </h2>
        </div>

        <div style="
            padding:0 4px 24px;
            font-family:Arial, Helvetica, sans-serif;
            color:#1c2b3a;
            font-size:14px;
            line-height:1.6;
        ">

            <p>
                Dear {escape_html(doc.customer or "Sir/Madam")},
            </p>

            <p>
                Greetings from NMTG Mechtrans Techniques Pvt. Ltd.
            </p>

            <p>
                With reference to Technical Evaluation
                <b>{escape_html(doc.name)}</b>, we require certain
                clarifications from your end to proceed further with our
                technical assessment. Kindly find the list of queries below.
            </p>

            <p>
                Kindly review the technical queries mentioned below and provide
                the required information at your earliest convenience.
            </p>

            <table style="
                border-collapse:collapse;
                width:100%;
                margin:16px 0;
                font-family:Arial, Helvetica, sans-serif;
                font-size:13px;
            ">

                <thead>
                    <tr>

                        <th style="
                            padding:8px 10px;
                            border:1px solid #ddd;
                            background:#f2f2f2;
                            text-align:left;
                        ">
                            Request No
                        </th>

                        <th style="
                            padding:8px 10px;
                            border:1px solid #ddd;
                            background:#f2f2f2;
                            text-align:left;
                        ">
                            Customer Requirement/Brief
                        </th>

                        <th style="
                            padding:8px 10px;
                            border:1px solid #ddd;
                            background:#f2f2f2;
                            text-align:left;
                        ">
                            Questions
                        </th>

                    </tr>
                </thead>

                <tbody>
                    {rows_html}
                </tbody>

            </table>

            <table
                role="presentation"
                cellpadding="0"
                cellspacing="0"
                border="0"
                align="center"
                style="margin:24px auto;"
            >
                <tr>
                    <td
                        style="
                            background-color:#000000;
                            border-radius:4px;
                        "
                        align="center"
                    >
                        <a
                            href="{link}"
                            target="_blank"
                            style="
                                display:inline-block;
                                padding:12px 28px;
                                font-family:Arial, Helvetica, sans-serif;
                                font-size:14px;
                                font-weight:600;
                                letter-spacing:.2px;
                                color:#ffffff !important;
                                text-decoration:none !important;
                                border:1px solid #000000;
                                border-radius:4px;
                            "
                        >
                            Click here to answer the questionary
                        </a>
                    </td>
                </tr>
            </table>

            <p style="
                font-size:12.5px;
                color:#5a6b78;
            ">
                Should the button above not work, please copy and paste the
                following link into your browser:<br>

                <a
                    href="{link}"
                    style="color:#0c355a;"
                >
                    {link}
                </a>
            </p>

            <p>
                Should you have any queries regarding the above, please feel
                free to reach out to us.
            </p>

            <p>
                Best regards,<br>
                <b>NMTG Mechtrans Techniques Pvt. Ltd.</b>
            </p>

        </div>
    """

    frappe.sendmail(
        recipients=[doc.email],
        subject=_("Technical Evaluation Questionary - {0}").format(doc.name),
        message=message,
        reference_doctype=doc.doctype,
        reference_name=doc.name,
    )

    doc.add_comment(
        "Info",
        _("Questionary email sent to {0}").format(doc.email)
    )

    return {"sent": True}



@frappe.whitelist(allow_guest=True)
def get_technical_evaluation_for_customer(technical_evaluation):
	"""Guest-safe read. Returns only what the portal is allowed to show —
	never the raw document — so guests never need read permission on the
	whole 'Technical Evaluation' doctype."""

	if not frappe.db.exists("Technical Evaluation", technical_evaluation):
		frappe.throw(_("Technical Evaluation not found"), frappe.DoesNotExistError)

	doc = frappe.get_doc("Technical Evaluation", technical_evaluation)

	details = {f: doc.get(f) for f in DETAIL_FIELDS}

	items = [{f: row.get(f) for f in ITEM_FIELDS} for row in doc.items]

	questionary = [
		{
			"name": row.name,
			"request_no": row.request_no,
			"question": row.question,
			"answer_from_customer": row.answer_from_customer,
		}
		for row in doc.questionary_for_technical_evaluation
	]

	return {"details": details, "items": items, "questionary": questionary}


@frappe.whitelist(allow_guest=True)
def save_technical_evaluation_answers(technical_evaluation, answers):
	"""Guest-safe write. Updates ONLY answer_from_customer on matching
	child rows — never touches any other field on the parent or on
	Items."""

	if isinstance(answers, str):
		answers = json.loads(answers)

	if not frappe.db.exists("Technical Evaluation", technical_evaluation):
		frappe.throw(_("Technical Evaluation not found"), frappe.DoesNotExistError)

	doc = frappe.get_doc("Technical Evaluation", technical_evaluation)

	answer_map = {a.get("name"): a.get("answer_from_customer") for a in answers if a.get("name")}

	updated = 0
	for row in doc.questionary_for_technical_evaluation:
		if row.name in answer_map:
			frappe.db.set_value(
				row.doctype,
				row.name,
				"answer_from_customer",
				answer_map[row.name],
				update_modified=True,
			)
			updated += 1

	frappe.db.commit()

	return {"name": doc.name, "updated": updated}



@frappe.whitelist()
def generate_apqp_template(parent, item_row):
    te = frappe.get_doc("Technical Evaluation", parent)
    row = next(r for r in te.items if r.name == item_row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "APQP"

    # ---- Fonts & fills ----
    font_company = Font(name="Arial", size=12, bold=True)
    font_title = Font(name="Arial", size=11, bold=True)
    font_label = Font(name="Arial", size=9, bold=True)
    font_value = Font(name="Arial", size=9)
    font_header = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    fill_company = PatternFill("solid", fgColor="DCE6F1")
    fill_label = PatternFill("solid", fgColor="DCE6F1")
    fill_header = PatternFill("solid", fgColor="1F4E78")

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---- Column widths ----
    widths = {"A": 7, "B": 14, "C": 14, "D": 14, "E": 16, "F": 14, "G": 16, "H": 14, "I": 14, "J": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- Row 1-2: Company / Format / Doc info ----
    ws.merge_cells("A1:B2")
    ws["A1"] = "NMTG Mechtrans Techniques Pvt. Ltd."
    ws["A1"].font = font_company
    ws["A1"].alignment = center
    ws["A1"].fill = fill_company

    ws.merge_cells("C1:H1")
    ws["C1"] = "FORMAT - ENGINEERING"
    ws["C1"].font = font_title
    ws["C1"].alignment = center

    ws.merge_cells("C2:H2")
    ws["C2"] = "TITLE \u2013 ADVANCE PRODUCT QUALITY PLANING CHART"
    ws["C2"].font = font_title
    ws["C2"].alignment = center

    ws["I1"] = "Document No"
    ws["I1"].font = font_label
    ws["I1"].fill = fill_label
    ws["I1"].alignment = center
    ws["J1"] = te.name  # Document No = parent Technical Evaluation name
    ws["J1"].font = font_value
    ws["J1"].alignment = center

    ws["I2"] = "Rev. No. & Date"
    ws["I2"].font = font_label
    ws["I2"].fill = fill_label
    ws["I2"].alignment = center
    ws["J2"] = f"00 & {frappe.utils.today()}"
    ws["J2"].font = font_value
    ws["J2"].alignment = center

    # ---- Row 3: Customer / Part Name / Page ----
    ws["A3"] = "Customer"
    ws["A3"].font = font_label
    ws["A3"].fill = fill_label
    ws["A3"].alignment = center

    ws.merge_cells("B3:D3")
    ws["B3"] = te.customer or ""
    ws["B3"].font = font_value
    ws["B3"].alignment = left

    ws["E3"] = "Part Name :"
    ws["E3"].font = font_label
    ws["E3"].fill = fill_label
    ws["E3"].alignment = center

    ws.merge_cells("F3:H3")
    ws["F3"] = row.item_name or row.item_code or ""
    ws["F3"].font = font_value
    ws["F3"].alignment = left

    ws["I3"] = "PAGE"
    ws["I3"].font = font_label
    ws["I3"].fill = fill_label
    ws["I3"].alignment = center
    ws["J3"] = "1 of 1"
    ws["J3"].font = font_value
    ws["J3"].alignment = center

    # ---- Row 4: Table headers ----
    ws.merge_cells("B4:D4")  # Activities Description spans 3 columns
    headers = {
        "A4": "Sr. No.",
        "B4": "Activities Description",
        "E4": "Status / action",
        "F4": "Reference",
        "G4": "Responsibility",
        "H4": "Target Date",
        "I4": "Actual Date",
        "J4": "Verification Status",
    }
    for cell, text in headers.items():
        c = ws[cell]
        c.value = text
        c.font = font_header
        c.fill = fill_header
        c.alignment = center

    # Borders for header block (rows 1-4)
    for r_ in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=10):
        for cell in r_:
            cell.border = border_all

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 30

    # ---- Blank rows for filling in (5-24) ----
    for r in range(5, 25):
        ws.merge_cells(f"B{r}:D{r}")  # keep Activities Description as one wide column, not 3
        for col_idx in range(1, 11):
            col = get_column_letter(col_idx)
            c = ws[f"{col}{r}"]
            c.border = border_all
            c.alignment = left if col == "B" else center

    # ---- Comments section ----
    comments_row = 26
    ws.merge_cells(f"A{comments_row}:B{comments_row + 1}")
    c = ws[f"A{comments_row}"]
    c.value = "Comments ( If any ) :"
    c.font = font_label
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells(f"C{comments_row}:J{comments_row + 1}")
    for r_ in ws.iter_rows(min_row=comments_row, max_row=comments_row + 1, min_col=1, max_col=10):
        for cell in r_:
            cell.border = border_all
    ws.row_dimensions[comments_row].height = 20
    ws.row_dimensions[comments_row + 1].height = 20

    # ---- Prepared By / Reviewed By ----
    sign_row = comments_row + 3
    ws.merge_cells(f"A{sign_row}:B{sign_row}")
    c = ws[f"A{sign_row}"]
    c.value = "Prepared By :"
    c.font = font_label
    c.alignment = left

    ws.merge_cells(f"C{sign_row}:E{sign_row}")

    ws.merge_cells(f"F{sign_row}:G{sign_row}")
    c = ws[f"F{sign_row}"]
    c.value = "Reviewed By :"
    c.font = font_label
    c.alignment = left

    ws.merge_cells(f"H{sign_row}:J{sign_row}")

    for col in range(1, 11):
        letter = get_column_letter(col)
        ws[f"{letter}{sign_row}"].border = Border(bottom=thin)
    ws.row_dimensions[sign_row].height = 22

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    # ---- Save & attach ----
    safe_te_name = te.name.replace("/", "-")
    file_path = f"/tmp/APQP_{safe_te_name}_{row.item_code}.xlsx"
    wb.save(file_path)

    with open(file_path, "rb") as f:
        content = f.read()

    file_doc = save_file(
        f"APQP_{safe_te_name}_{row.item_code}.xlsx", content,
        "Technical Evaluation", te.name, is_private=1
    )
    return file_doc.file_url


def get_prospect_recipients(doc):
	lead_names = [row.lead for row in (doc.get("leads") or []) if row.lead]
	if not lead_names:
		return None, []

	emails = []
	primary_email = None

	for lead_name in lead_names:
		if frappe.db.get_value("Lead", lead_name, "workflow_state") != "Qualified":
			continue

		lead_doc = frappe.get_doc("Lead", lead_name)
		for row in (lead_doc.get("custom_contact_info") or []):
			if not row.email_id:
				continue
			if row.primary_contact and not primary_email:
				primary_email = row.email_id
			if row.email_id not in emails:
				emails.append(row.email_id)

	if not emails:
		return None, []

	to = primary_email or emails[-1]
	cc = [e for e in emails if e != to]

	return to, cc


@frappe.whitelist()
def send_prospect_stage_email(prospect, stage_key, subject, message):
	if stage_key not in STAGE_FIELD_MAP:
		frappe.throw(_("Invalid stage key: {0}").format(stage_key))

	if not frappe.has_permission("Prospect", "write", prospect):
		frappe.throw(_("Not permitted"), frappe.PermissionError)

	doc = frappe.get_doc("Prospect", prospect)

	to, cc = get_prospect_recipients(doc)
	if not to:
		frappe.throw(_("No contact email addresses found to send to."))

	frappe.sendmail(
		recipients=[to],
		cc=cc,
		subject=subject,
		message=message,
		reference_doctype="Prospect",
		reference_name=prospect,
		attachments=get_email_attachments(),
	)

	field_key = STAGE_FIELD_MAP[stage_key]
	sent_on = frappe.utils.now()
	doc.db_set(f"custom_{field_key}_email_sent", 1, update_modified=False)
	doc.db_set(f"custom_{field_key}_email_sent_on", sent_on, update_modified=False)

	return {"sent_on": sent_on, "recipient": to, "cc": cc}



def set_custom_dealer(doc, method=None):
    
    if not doc.is_new():
        return

    if not doc.customer_name:
        doc.custom_dealer = ""
        return

    customer_name = doc.customer_name.strip().lower()

    rows = frappe.get_all(
        "Dealer Customer CT",
        fields=["parent", "dealer_customer"],
        ignore_permissions=True,
    )

    matched_parent = ""
    for row in rows:
        if row.dealer_customer and row.dealer_customer.strip().lower() == customer_name:
            matched_parent = row.parent
            break

    doc.custom_dealer = matched_parent

    if matched_parent:
        frappe.msgprint(
            msg=_("This customer is a Dealer Customer of <b>{0}</b>").format(matched_parent),
            title=_("Dealer Customer"),
            indicator="blue",
            alert=True,
        )



@frappe.whitelist(allow_guest=True)
def get_supplier_scope(supplier):
    if not supplier:
        frappe.throw("Supplier is required")

    scope = frappe.db.get_value("Supplier", supplier, "custom_supplier_scope")
    return scope


@frappe.whitelist(allow_guest=True)
def get_csrf_token():
   
    token = frappe.sessions.get_csrf_token()
    frappe.db.commit() 
    return token


def get_formatted_size(required_fields_json, doctype="Item", separator="<br>"):
    if not required_fields_json:
        return ""

    try:
        data = json.loads(required_fields_json)
    except Exception:
        return ""

    meta = frappe.get_meta(doctype)
    lines = []
    for key, value in data.items():
        df = meta.get_field(key)
        label = df.label if df else key.replace("custom_", "").replace("_", " ").title()
        lines.append(f"{label}: {value}")

    return separator.join(lines)